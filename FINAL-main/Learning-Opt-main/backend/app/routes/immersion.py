import os
from flask import Blueprint, request, send_file, jsonify
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
from io import BytesIO
import json

immersion_bp = Blueprint('immersion', __name__)

basedir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
TEMPLATE_PATH = os.path.join(basedir, "uploads", "templates", "Grades.xlsx")
UPLOAD_JSON_PATH = os.path.join(basedir, "backend", "app", "static", "excel", "uploaded_data.json")

print("Resolved TEMPLATE_PATH:", TEMPLATE_PATH)
print("Exists?", os.path.exists(TEMPLATE_PATH))

@immersion_bp.route("/fill-template", methods=["POST"])
def fill_template():
    if "file" not in request.files:
        return jsonify({"error": "No file part in the request"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No selected file"}), 400

    try:
        file.stream.seek(0)
        wb_uploaded = load_workbook(file.stream)
        ws_uploaded = wb_uploaded.active

        data = []
        # Assuming your Excel columns are aligned: Last, First, Middle, Strand, Dept, plus scores columns
        for row in ws_uploaded.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            data.append({
                "LAST NAME": row[0] or "",
                "FIRST NAME": row[1] or "",
                "MIDDLE NAME": row[2] or "",
                "STRAND": row[3] or "",
                "DEPARTMENT": row[4] or "",
                # Add other keys if needed (e.g., W1, CO, etc.) here based on column indexes
            })

        # Save data as JSON for frontend consumption
        os.makedirs(os.path.dirname(UPLOAD_JSON_PATH), exist_ok=True)
        with open(UPLOAD_JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)

        # Also generate the filled Excel template
        if not os.path.exists(TEMPLATE_PATH):
            return jsonify({"error": f"Template not found at {TEMPLATE_PATH}"}), 500

        wb_template = load_workbook(TEMPLATE_PATH)
        ws_template = wb_template.active

        start_row = 10
        max_columns = 10  # Adjust as needed for formatting copy

        # Copy formatting rows if needed
        preformatted_rows = [
            row for row in range(start_row, ws_template.max_row + 1)
            if ws_template.cell(row=row, column=2).value or ws_template.cell(row=row, column=2).has_style
        ]
        template_rows = len(preformatted_rows)
        extra_rows_needed = len(data) - template_rows

        if extra_rows_needed > 0:
            last_template_row = preformatted_rows[-1]
            for i in range(1, extra_rows_needed + 1):
                source_row_idx = last_template_row
                target_row_idx = last_template_row + i

                for col in range(1, max_columns + 1):
                    source_cell = ws_template.cell(row=source_row_idx, column=col)
                    target_cell = ws_template.cell(row=target_row_idx, column=col)

                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.border = copy(source_cell.border)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.number_format = copy(source_cell.number_format)
                        target_cell.protection = copy(source_cell.protection)
                        target_cell.alignment = copy(source_cell.alignment)

        # Fill template with data
        for idx, entry in enumerate(data):
            row = start_row + idx
            ws_template.cell(row=row, column=1, value=idx + 1)               # #
            ws_template.cell(row=row, column=2, value=entry["LAST NAME"])    # B
            ws_template.cell(row=row, column=3, value=entry["FIRST NAME"])   # C
            ws_template.cell(row=row, column=4, value=entry["MIDDLE NAME"])  # D
            ws_template.cell(row=row, column=5, value=entry["STRAND"])       # E
            ws_template.cell(row=row, column=6, value=entry["DEPARTMENT"])   # F

        # Auto adjust columns A-F
        for col in range(1, 7):
            max_length = 0
            col_letter = get_column_letter(col)
            for row in range(start_row, start_row + len(data)):
                cell = ws_template.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_template.column_dimensions[col_letter].width = max_length + 2

        # Save filled template to memory and send back
        output = BytesIO()
        wb_template.save(output)
        output.seek(0)

        # Return both JSON data and file download link
        # Here we just return JSON data for frontend table; you can create a separate endpoint for file download
        return jsonify({
            "message": "Data extracted and template filled successfully",
            "rows": data
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@immersion_bp.route("/data", methods=["GET"])
def get_immersion_data():
    if not os.path.exists(UPLOAD_JSON_PATH):
        return jsonify({"rows": []})

    with open(UPLOAD_JSON_PATH, "r", encoding="utf-8") as f:
        rows = json.load(f)

    return jsonify({"rows": rows})
